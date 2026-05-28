import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="26AS Reconciliation Assistant",
    page_icon="⚖️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* Background */
.stApp {
    background: #F7F4EF;
}

/* Hide default streamlit chrome */
#MainMenu, footer, header { visibility: hidden; }

/* Hero */
.hero {
    background: linear-gradient(135deg, #1F3864 0%, #2E75B6 100%);
    border-radius: 16px;
    padding: 48px 40px 40px;
    margin-bottom: 32px;
    position: relative;
    overflow: hidden;
}
.hero::before {
    content: "26AS";
    position: absolute;
    right: -10px;
    top: -20px;
    font-family: 'DM Serif Display', serif;
    font-size: 160px;
    color: rgba(255,255,255,0.05);
    pointer-events: none;
}
.hero h1 {
    font-family: 'DM Serif Display', serif;
    font-size: 2.6rem;
    color: #FFFFFF;
    margin: 0 0 8px 0;
    line-height: 1.15;
}
.hero p {
    color: rgba(255,255,255,0.75);
    font-size: 1rem;
    margin: 0;
    font-weight: 300;
}
.hero .badge {
    display: inline-block;
    background: rgba(255,255,255,0.15);
    border: 1px solid rgba(255,255,255,0.25);
    color: white;
    padding: 4px 12px;
    border-radius: 20px;
    font-size: 0.78rem;
    margin-bottom: 16px;
    letter-spacing: 0.5px;
}

/* Upload cards */
.upload-card {
    background: white;
    border-radius: 12px;
    padding: 28px 24px;
    border: 2px dashed #C5D8EE;
    text-align: center;
    transition: border-color 0.2s;
}
.upload-card:hover { border-color: #2E75B6; }
.upload-label {
    font-family: 'DM Serif Display', serif;
    font-size: 1.1rem;
    color: #1F3864;
    margin-bottom: 6px;
}
.upload-sub {
    font-size: 0.82rem;
    color: #7A8CA0;
}

/* KPI cards */
.kpi-row { display: flex; gap: 16px; margin: 24px 0; flex-wrap: wrap; }
.kpi {
    flex: 1;
    min-width: 140px;
    background: white;
    border-radius: 12px;
    padding: 20px 18px;
    box-shadow: 0 1px 8px rgba(0,0,0,0.06);
}
.kpi-label { font-size: 0.78rem; color: #7A8CA0; text-transform: uppercase; letter-spacing: 0.8px; margin-bottom: 6px; }
.kpi-value { font-family: 'DM Serif Display', serif; font-size: 1.8rem; color: #1F3864; }
.kpi-sub { font-size: 0.78rem; color: #7A8CA0; margin-top: 2px; }
.kpi.green .kpi-value { color: #166534; }
.kpi.red .kpi-value { color: #C00000; }
.kpi.amber .kpi-value { color: #92400E; }

/* Section titles */
.section-title {
    font-family: 'DM Serif Display', serif;
    font-size: 1.4rem;
    color: #1F3864;
    margin: 32px 0 16px;
    padding-bottom: 8px;
    border-bottom: 2px solid #D6E4F0;
}

/* Mismatch badge */
.badge-matched    { background:#D1FAE5; color:#065F46; padding:3px 10px; border-radius:20px; font-size:0.78rem; font-weight:600; }
.badge-26only     { background:#FCE4D6; color:#C00000; padding:3px 10px; border-radius:20px; font-size:0.78rem; font-weight:600; }
.badge-booksonly  { background:#FCE4D6; color:#C00000; padding:3px 10px; border-radius:20px; font-size:0.78rem; font-weight:600; }
.badge-section    { background:#FEF3C7; color:#92400E; padding:3px 10px; border-radius:20px; font-size:0.78rem; font-weight:600; }
.badge-amount     { background:#FEF3C7; color:#92400E; padding:3px 10px; border-radius:20px; font-size:0.78rem; font-weight:600; }
.badge-tan        { background:#FFEDD5; color:#7C2D12; padding:3px 10px; border-radius:20px; font-size:0.78rem; font-weight:600; }

/* Action box */
.action-box {
    background: #F0F7FF;
    border-left: 3px solid #2E75B6;
    padding: 8px 12px;
    border-radius: 0 6px 6px 0;
    font-size: 0.82rem;
    color: #1F3864;
    margin-top: 4px;
}

/* Download button */
.stDownloadButton > button {
    background: #1F3864 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 12px 28px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.95rem !important;
    font-weight: 500 !important;
    letter-spacing: 0.3px !important;
}
.stDownloadButton > button:hover { background: #2E75B6 !important; }

/* Disclaimer */
.disclaimer {
    background: #FFFBEB;
    border: 1px solid #FCD34D;
    border-radius: 8px;
    padding: 14px 18px;
    font-size: 0.82rem;
    color: #78350F;
    margin-top: 32px;
}

/* Recon button */
.stButton > button {
    background: linear-gradient(135deg, #1F3864, #2E75B6) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 14px 40px !important;
    font-size: 1rem !important;
    font-weight: 500 !important;
    font-family: 'DM Sans', sans-serif !important;
    width: 100%;
    margin-top: 12px;
}

/* Step indicator */
.step { display:flex; align-items:center; gap:10px; margin-bottom:8px; }
.step-num {
    width:28px; height:28px; border-radius:50%;
    background:#1F3864; color:white;
    display:flex; align-items:center; justify-content:center;
    font-size:0.8rem; font-weight:600; flex-shrink:0;
}
.step-text { font-size:0.88rem; color:#4A5568; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

AMT_TOLERANCE = 0.02

def load_26as(file) -> pd.DataFrame:
    df = pd.read_excel(file, sheet_name=0, skiprows=3, header=0)
    df.columns = df.columns.str.strip().str.replace("\n", " ")
    df = df.rename(columns={
        df.columns[0]: "Sr", df.columns[1]: "TAN_26AS",
        df.columns[2]: "Vendor_26AS", df.columns[3]: "Section_26AS",
        df.columns[4]: "Date_26AS", df.columns[5]: "Amount_26AS",
        df.columns[6]: "TDS_26AS", df.columns[7]: "Deposited_26AS",
        df.columns[8]: "Quarter_26AS", df.columns[9]: "Category_26AS",
    })
    df = df[df["Sr"].apply(lambda x: str(x).strip().replace(".0","").isdigit())].copy()
    df["TDS_26AS"] = pd.to_numeric(df["TDS_26AS"], errors="coerce")
    df["Amount_26AS"] = pd.to_numeric(df["Amount_26AS"], errors="coerce")
    return df.reset_index(drop=True)

def load_books(file) -> pd.DataFrame:
    df = pd.read_excel(file, sheet_name=0, skiprows=3, header=0)
    df.columns = df.columns.str.strip().str.replace("\n", " ")
    df = df.rename(columns={
        df.columns[0]: "Sr", df.columns[1]: "Vendor_Books",
        df.columns[2]: "TAN_Books", df.columns[3]: "Section_Books",
        df.columns[4]: "Date_Books", df.columns[5]: "InvAmt_Books",
        df.columns[6]: "TDS_Books", df.columns[7]: "Quarter_Books",
        df.columns[8]: "TDS_Claimed", df.columns[9]: "Category_Books",
        df.columns[10]: "Remarks",
    })
    df = df[df["Sr"].apply(lambda x: str(x).strip().replace(".0","").isdigit())].copy()
    df["TDS_Books"] = pd.to_numeric(df["TDS_Books"], errors="coerce")
    return df.reset_index(drop=True)

def reconcile(df26: pd.DataFrame, dfb: pd.DataFrame) -> pd.DataFrame:
    matched_26as, matched_books = set(), set()
    results = []

    for i26, r26 in df26.iterrows():
        tan26  = str(r26["TAN_26AS"]).strip()
        amt26  = r26["TDS_26AS"]
        sec26  = str(r26["Section_26AS"]).strip()
        vend26 = str(r26["Vendor_26AS"]).strip()
        qtr26  = str(r26["Quarter_26AS"]).strip()

        best_match, best_score, best_type = None, 0, None

        for ib, rb in dfb.iterrows():
            if ib in matched_books: continue
            tanb  = str(rb["TAN_Books"]).strip()
            amtb  = rb["TDS_Books"]
            secb  = str(rb["Section_Books"]).strip()
            vendb = str(rb["Vendor_Books"]).strip()
            name_score = fuzz.token_sort_ratio(vend26, vendb)

            if tan26 == tanb and abs(amt26 - amtb) <= 1:
                mtype, score = ("Matched", 100) if sec26 == secb else ("Section Mismatch", 95)
            elif name_score >= 85 and tan26 != tanb and abs(amt26 - amtb) <= 1:
                mtype, score = "TAN Mismatch", 90
            elif tan26 == tanb and abs(amt26 - amtb) / max(amt26, 1) <= AMT_TOLERANCE:
                mtype, score = "Amount Mismatch", 85
            else:
                continue

            if score > best_score:
                best_score, best_match, best_type = score, (ib, rb), mtype

        if best_match:
            ib, rb = best_match
            matched_26as.add(i26)
            matched_books.add(ib)
            results.append({
                "Vendor": r26["Vendor_26AS"],
                "TAN (26AS)": tan26, "TAN (Books)": str(rb["TAN_Books"]).strip(),
                "Section (26AS)": sec26, "Section (Books)": str(rb["Section_Books"]).strip(),
                "TDS in 26AS": amt26, "TDS in Books": rb["TDS_Books"],
                "Difference": round(amt26 - rb["TDS_Books"], 0),
                "Quarter": qtr26, "Mismatch Type": best_type,
                "Action": rb["Remarks"] if str(rb["Remarks"]).strip() not in ("", "nan") else "None — fully reconciled.",
            })
        else:
            results.append({
                "Vendor": vend26, "TAN (26AS)": tan26, "TAN (Books)": "—",
                "Section (26AS)": sec26, "Section (Books)": "—",
                "TDS in 26AS": amt26, "TDS in Books": 0,
                "Difference": amt26, "Quarter": qtr26,
                "Mismatch Type": "26AS Only — Not in Books",
                "Action": "Invoice not yet accounted in books. Common where March invoice is booked in April by the assessee. Verify with party and accrue TDS credit in the correct period.",
            })

    for ib, rb in dfb.iterrows():
        if ib not in matched_books:
            results.append({
                "Vendor": rb["Vendor_Books"], "TAN (26AS)": "—",
                "TAN (Books)": str(rb["TAN_Books"]).strip(),
                "Section (26AS)": "—", "Section (Books)": str(rb["Section_Books"]).strip(),
                "TDS in 26AS": 0, "TDS in Books": rb["TDS_Books"],
                "Difference": -rb["TDS_Books"], "Quarter": str(rb["Quarter_Books"]).strip(),
                "Mismatch Type": "Books Only — Not in 26AS",
                "Action": str(rb["Remarks"]).strip() if str(rb["Remarks"]).strip() not in ("", "nan")
                          else "TDS deducted and booked but not yet in 26AS. Follow up with deductor to file / revise TDS return. Obtain Form 16A after filing.",
            })

    df_res = pd.DataFrame(results)
    order = {"26AS Only — Not in Books": 0, "Books Only — Not in 26AS": 1,
             "Section Mismatch": 2, "Amount Mismatch": 3, "TAN Mismatch": 4, "Matched": 5}
    df_res["_ord"] = df_res["Mismatch Type"].map(order).fillna(9)
    return df_res.sort_values("_ord").drop(columns="_ord").reset_index(drop=True)

def badge_html(mtype):
    cls_map = {
        "Matched": "matched", "26AS Only — Not in Books": "26only",
        "Books Only — Not in 26AS": "booksonly", "Section Mismatch": "section",
        "Amount Mismatch": "amount", "TAN Mismatch": "tan",
    }
    label_map = {
        "Matched": "✅ Matched", "26AS Only — Not in Books": "🔴 26AS Only",
        "Books Only — Not in 26AS": "🔴 Books Only", "Section Mismatch": "🟡 Section Mismatch",
        "Amount Mismatch": "🟡 Amount Mismatch", "TAN Mismatch": "🟠 TAN Mismatch",
    }
    cls   = cls_map.get(mtype, "matched")
    label = label_map.get(mtype, mtype)
    return f'<span class="badge-{cls}">{label}</span>'

def inr(v):
    try: return f"₹{int(v):,}"
    except: return "₹0"

def build_excel(df_res: pd.DataFrame, company: str) -> bytes:
    DARK_BLUE="1F3864"; MID_BLUE="2E75B6"; LIGHT_BLUE="D6E4F0"
    GREEN="E2EFDA"; GREEN_FG="166534"; RED_BG="FCE4D6"; RED_FG="C00000"
    AMBER_BG="FFF2CC"; AMBER_FG="7F6000"; TAN_BG="F4CCAC"; TAN_FG="843C0C"
    WHITE="FFFFFF"; GREY="F5F5F5"; INR_FMT="#,##0"

    thin = Side(style="thin", color="BFBFBF")
    def bdr(ws,r,c): ws.cell(r,c).border = Border(left=thin,right=thin,top=thin,bottom=thin)
    def hc(ws,r,c,v,bg=DARK_BLUE,fg=WHITE,wrap=False):
        cell=ws.cell(r,c,v); cell.font=Font(name="Arial",bold=True,color=fg,size=10)
        cell.fill=PatternFill("solid",start_color=bg); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=wrap)
    def dc(ws,r,c,v,bg=WHITE,fg="000000",bold=False,fmt=None,align="left"):
        cell=ws.cell(r,c,v); cell.font=Font(name="Arial",size=10,color=fg,bold=bold)
        cell.fill=PatternFill("solid",start_color=bg); cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
        if fmt: cell.number_format=fmt

    def row_style(mtype):
        m={"Matched":(GREEN,GREEN_FG),"26AS Only — Not in Books":(RED_BG,RED_FG),
           "Books Only — Not in 26AS":(RED_BG,RED_FG),"Section Mismatch":(AMBER_BG,AMBER_FG),
           "Amount Mismatch":(AMBER_BG,AMBER_FG),"TAN Mismatch":(TAN_BG,TAN_FG)}
        return m.get(mtype,(WHITE,"000000"))

    wb = Workbook()

    # Sheet 1 — Detail
    ws1 = wb.active; ws1.title = "Detailed Reconciliation"
    ws1.freeze_panes = "A5"; ws1.sheet_view.zoomScale = 85
    ws1.merge_cells("A1:K1")
    ws1["A1"].value = f"26AS RECONCILIATION REPORT — FY 2024-25 | {company}"
    ws1["A1"].font = Font(name="Arial",bold=True,size=12,color=WHITE)
    ws1["A1"].fill = PatternFill("solid",start_color=DARK_BLUE)
    ws1["A1"].alignment = Alignment(horizontal="center",vertical="center")
    ws1.row_dimensions[1].height = 26
    ws1.merge_cells("A2:K2")
    ws1["A2"].value = "AI-assisted reconciliation | Matching: Exact TAN+Amount, Fuzzy Vendor, ±2% Tolerance | Mismatches sorted first"
    ws1["A2"].font = Font(name="Arial",size=9,italic=True,color=DARK_BLUE)
    ws1["A2"].fill = PatternFill("solid",start_color=LIGHT_BLUE)
    ws1["A2"].alignment = Alignment(horizontal="center",vertical="center")
    ws1.row_dimensions[2].height = 15
    ws1.merge_cells("A3:K3")
    ws1["A3"].value = "GREEN = Matched  |  RED = Only in one source  |  AMBER = Section/Amount mismatch  |  ORANGE = TAN mismatch"
    ws1["A3"].font = Font(name="Arial",size=9,color=WHITE)
    ws1["A3"].fill = PatternFill("solid",start_color=MID_BLUE)
    ws1["A3"].alignment = Alignment(horizontal="center",vertical="center")
    ws1.row_dimensions[3].height = 15

    hdrs = ["Vendor Name","TAN (26AS)","TAN (Books)","Section\n(26AS)","Section\n(Books)",
            "TDS in\n26AS (₹)","TDS in\nBooks (₹)","Difference\n(₹)","Quarter","Mismatch Type","Action Required"]
    widths = [32,14,14,11,11,15,15,14,9,26,52]
    for i,(h,w) in enumerate(zip(hdrs,widths),1):
        hc(ws1,4,i,h,wrap=True); ws1.column_dimensions[get_column_letter(i)].width=w
    ws1.row_dimensions[4].height = 36

    for ri,row in df_res.iterrows():
        mtype=row["Mismatch Type"]; bg,fg=row_style(mtype); er=ri+5
        fields=[(row["Vendor"],"left"),(row["TAN (26AS)"],"center"),(row["TAN (Books)"],"center"),
                (row["Section (26AS)"],"center"),(row["Section (Books)"],"center"),
                (row["TDS in 26AS"],"right"),(row["TDS in Books"],"right"),(row["Difference"],"right"),
                (row["Quarter"],"center"),(mtype,"center"),(row["Action"],"left")]
        for ci,(val,align) in enumerate(fields,1):
            fmt=INR_FMT if ci in (6,7,8) else None
            dc(ws1,er,ci,val,bg=bg,fg=fg,bold=(ci==10 and mtype!="Matched"),fmt=fmt,align=align)
            bdr(ws1,er,ci)
        ws1.row_dimensions[er].height=42

    tr=len(df_res)+5
    ws1.merge_cells(f"A{tr}:E{tr}")
    tc=ws1[f"A{tr}"]; tc.value="TOTALS"; tc.font=Font(name="Arial",bold=True,size=10,color=WHITE)
    tc.fill=PatternFill("solid",start_color=DARK_BLUE); tc.alignment=Alignment(horizontal="right",vertical="center")
    for cn,cl in [(6,"F"),(7,"G"),(8,"H")]:
        c=ws1.cell(tr,cn,f"=SUM({cl}5:{cl}{tr-1})")
        c.font=Font(name="Arial",bold=True,color=WHITE,size=10)
        c.fill=PatternFill("solid",start_color=DARK_BLUE)
        c.number_format=INR_FMT; c.alignment=Alignment(horizontal="right",vertical="center")

    # Sheet 2 — Summary
    ws2 = wb.create_sheet("Executive Summary")
    ws2.sheet_view.zoomScale=90
    for i,w in enumerate([36,18,22,22,18,46],1):
        ws2.column_dimensions[get_column_letter(i)].width=w

    ws2.merge_cells("A1:F1")
    ws2["A1"].value=f"EXECUTIVE SUMMARY — 26AS Reconciliation | FY 2024-25 | {company}"
    ws2["A1"].font=Font(name="Arial",bold=True,size=13,color=WHITE)
    ws2["A1"].fill=PatternFill("solid",start_color=DARK_BLUE)
    ws2["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[1].height=28

    # KPIs
    hc(ws2,3,1,"Metric",bg=MID_BLUE); hc(ws2,3,2,"Value",bg=MID_BLUE)
    ws2.row_dimensions[3].height=20
    total26=df_res["TDS in 26AS"].sum(); totalb=df_res["TDS in Books"].sum()
    kpis=[("Total entries reconciled",len(df_res)),("Entries matched",(df_res["Mismatch Type"]=="Matched").sum()),
          ("Entries with mismatches",(df_res["Mismatch Type"]!="Matched").sum()),
          ("Total TDS in 26AS (₹)",int(total26)),("Total TDS in Books (₹)",int(totalb)),
          ("Net difference (₹)",int(total26-totalb))]
    for ri,(label,val) in enumerate(kpis,4):
        dc(ws2,ri,1,label,bg=GREY if ri%2==0 else WHITE)
        fmt=INR_FMT if "₹" in label else None
        dc(ws2,ri,2,val,bg=GREY if ri%2==0 else WHITE,align="right",fmt=fmt)
        ws2.row_dimensions[ri].height=20

    # Category table
    hdr_r=11
    for i,h in enumerate(["Mismatch Category","Entries","TDS 26AS (₹)","TDS Books (₹)","Difference (₹)","Recommended Action"],1):
        hc(ws2,hdr_r,i,h,wrap=True)
    ws2.row_dimensions[hdr_r].height=28

    cats=["Matched","26AS Only — Not in Books","Books Only — Not in 26AS","Section Mismatch","Amount Mismatch","TAN Mismatch"]
    actions={"Matched":"No action required — fully reconciled.",
             "26AS Only — Not in Books":"Book invoice; accrue TDS credit in correct period.",
             "Books Only — Not in 26AS":"Follow up with deductor for TDS return filing; obtain Form 16A.",
             "Section Mismatch":"Verify nature of service; reclassify; assess rate differential.",
             "Amount Mismatch":"Confirm correct amount with deductor; raise debit/credit note.",
             "TAN Mismatch":"Update vendor master; verify 26AS credit linkage with new TAN."}
    for ri,cat in enumerate(cats,hdr_r+1):
        sub=df_res[df_res["Mismatch Type"]==cat]
        if len(sub)==0: continue
        bg,fg=row_style(cat); bold=(cat!="Matched")
        dc(ws2,ri,1,cat,bg=bg,fg=fg,bold=bold)
        dc(ws2,ri,2,len(sub),bg=bg,fg=fg,align="center")
        dc(ws2,ri,3,int(sub["TDS in 26AS"].sum()),bg=bg,fg=fg,fmt=INR_FMT,align="right")
        dc(ws2,ri,4,int(sub["TDS in Books"].sum()),bg=bg,fg=fg,fmt=INR_FMT,align="right")
        dc(ws2,ri,5,int(sub["Difference"].sum()),bg=bg,fg=fg,fmt=INR_FMT,align="right")
        dc(ws2,ri,6,actions.get(cat,""),bg=bg,fg=fg)
        for c in range(1,7): bdr(ws2,ri,c)
        ws2.row_dimensions[ri].height=26

    # Grand total
    gt=hdr_r+len(cats)+1
    for ci,(v,al) in enumerate([(company,),(len(df_res),),(int(total26),),(int(totalb),),(int(total26-totalb),),("",)],1):
        dc(ws2,gt,ci,v[0],bg=DARK_BLUE,fg=WHITE,bold=True,
           fmt=INR_FMT if ci in(3,4,5) else None,align="right" if ci>1 else "left")
        bdr(ws2,gt,ci)
    ws2.cell(gt,1).value="GRAND TOTAL"
    ws2.row_dimensions[gt].height=24

    # Disclaimer
    disc_r=gt+2
    ws2.merge_cells(f"A{disc_r}:F{disc_r+4}")
    ws2[f"A{disc_r}"].value=(
        "DISCLAIMER: This reconciliation is based on the data files uploaded and is indicative only. "
        "Matching logic uses exact TAN+amount, fuzzy vendor name (≥85% similarity), and ±2% amount tolerance. "
        "Always verify against the actual TRACES 26AS download and audited books before filing returns or responding to notices. "
        "This tool does not constitute professional tax advice.")
    ws2[f"A{disc_r}"].font=Font(name="Arial",size=9,italic=True,color="78350F")
    ws2[f"A{disc_r}"].fill=PatternFill("solid",start_color="FFFBEB")
    ws2[f"A{disc_r}"].alignment=Alignment(wrap_text=True,vertical="top")
    ws2.row_dimensions[disc_r].height=72

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="hero">
    <div class="badge">AI-ASSISTED · FY 2024-25</div>
    <h1>26AS Reconciliation<br>Assistant</h1>
    <p>Upload your 26AS extract and books ledger. Get a colour-coded reconciliation<br>
    with mismatch detection, root-cause notes, and a downloadable report — in seconds.</p>
</div>
""", unsafe_allow_html=True)

# ── How it works ──────────────────────────────────────────────────────────────
with st.expander("How it works", expanded=False):
    st.markdown("""
    <div style="padding: 4px 0;">
        <div class="step"><div class="step-num">1</div><div class="step-text">Upload your <b>26AS Excel extract</b> (Part I — TDS on payments) downloaded from TRACES.</div></div>
        <div class="step"><div class="step-num">2</div><div class="step-text">Upload your <b>TDS Receivable Ledger</b> from books (Tally, Zoho, SAP, or custom Excel).</div></div>
        <div class="step"><div class="step-num">3</div><div class="step-text">The engine matches entries using <b>exact TAN + amount</b>, <b>fuzzy vendor name</b> (for TAN mismatches), and <b>±2% amount tolerance</b>.</div></div>
        <div class="step"><div class="step-num">4</div><div class="step-text">Every mismatch is <b>categorised and explained</b> with a CA-written action note. Download the full report.</div></div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("""
    **Detects 5 mismatch types:**
    - 🔴 Present in 26AS, absent in books
    - 🔴 Present in books, absent in 26AS (Q4 filing lag)
    - 🟡 Section mismatch (e.g. 194C booked vs 194J in 26AS)
    - 🟡 Amount mismatch (within 2% tolerance)
    - 🟠 TAN mismatch (same vendor, different TAN — group companies / restructuring)
    """)

st.markdown("---")

# ── File uploads ──────────────────────────────────────────────────────────────
col1, col2, col3 = st.columns([1, 1, 1])

with col1:
    st.markdown('<div class="upload-label">📄 26AS Extract</div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-sub">Excel file from TRACES (Part I)</div>', unsafe_allow_html=True)
    file_26as = st.file_uploader("Upload 26AS", type=["xlsx"], key="f26", label_visibility="collapsed")

with col2:
    st.markdown('<div class="upload-label">📒 Books Ledger</div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-sub">TDS Receivable ledger from your accounting system</div>', unsafe_allow_html=True)
    file_books = st.file_uploader("Upload Books", type=["xlsx"], key="fb", label_visibility="collapsed")

with col3:
    st.markdown('<div class="upload-label">🏢 Company Name</div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-sub">For report header</div>', unsafe_allow_html=True)
    company = st.text_input("Company", value="Mehta Exports Pvt Ltd", label_visibility="collapsed")
    run = st.button("⚖️ Run Reconciliation")

# ── Run ───────────────────────────────────────────────────────────────────────
if run:
    if not file_26as or not file_books:
        st.warning("Please upload both files before running.")
    else:
        with st.spinner("Matching entries..."):
            try:
                df26 = load_26as(file_26as)
                dfb  = load_books(file_books)
                df_res = reconcile(df26, dfb)

                total_26as = df26["TDS_26AS"].sum()
                total_books = dfb["TDS_Books"].sum()
                matched_n  = (df_res["Mismatch Type"] == "Matched").sum()
                mismatch_n = (df_res["Mismatch Type"] != "Matched").sum()
                diff = df_res["Difference"].sum()

                # ── KPIs ──────────────────────────────────────────────────────
                st.markdown('<div class="section-title">Results at a Glance</div>', unsafe_allow_html=True)
                k1,k2,k3,k4,k5,k6 = st.columns(6)
                kpi_data = [
                    (k1, "Total Entries", len(df_res), "", ""),
                    (k2, "Matched", matched_n, f"{matched_n/len(df_res)*100:.0f}% of total", "green"),
                    (k3, "Mismatches", mismatch_n, "require action", "red" if mismatch_n > 0 else "green"),
                    (k4, "TDS in 26AS", inr(total_26as), "from TRACES", ""),
                    (k5, "TDS in Books", inr(total_books), "from ledger", ""),
                    (k6, "Net Difference", inr(abs(diff)), "26AS vs Books", "amber" if abs(diff) > 0 else "green"),
                ]
                for col, label, val, sub, cls in kpi_data:
                    with col:
                        st.markdown(f"""
                        <div class="kpi {cls}">
                            <div class="kpi-label">{label}</div>
                            <div class="kpi-value">{val}</div>
                            <div class="kpi-sub">{sub}</div>
                        </div>""", unsafe_allow_html=True)

                # ── Category summary ──────────────────────────────────────────
                st.markdown('<div class="section-title">Mismatch Breakdown</div>', unsafe_allow_html=True)
                cats = ["26AS Only — Not in Books","Books Only — Not in 26AS",
                        "Section Mismatch","Amount Mismatch","TAN Mismatch","Matched"]
                for cat in cats:
                    sub = df_res[df_res["Mismatch Type"] == cat]
                    if len(sub) == 0: continue
                    icon = "✅" if cat == "Matched" else "⚠️"
                    with st.expander(f"{icon}  {cat}  —  {len(sub)} entr{'y' if len(sub)==1 else 'ies'}  |  26AS: {inr(sub['TDS in 26AS'].sum())}  |  Books: {inr(sub['TDS in Books'].sum())}"):
                        for _, row in sub.iterrows():
                            st.markdown(f"""
                            <div style="background:white;border-radius:10px;padding:14px 18px;margin-bottom:10px;
                                        box-shadow:0 1px 6px rgba(0,0,0,0.06);">
                                <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:12px;">
                                    <div>
                                        <div style="font-weight:600;color:#1F3864;font-size:0.95rem;">{row['Vendor']}</div>
                                        <div style="font-size:0.8rem;color:#7A8CA0;margin-top:2px;">
                                            TAN 26AS: <b>{row['TAN (26AS)']}</b> &nbsp;|&nbsp;
                                            TAN Books: <b>{row['TAN (Books)']}</b> &nbsp;|&nbsp;
                                            Section 26AS: <b>{row['Section (26AS)']}</b> &nbsp;|&nbsp;
                                            Section Books: <b>{row['Section (Books)']}</b> &nbsp;|&nbsp;
                                            Quarter: <b>{row['Quarter']}</b>
                                        </div>
                                    </div>
                                    <div style="text-align:right;flex-shrink:0;">
                                        <div style="font-size:0.78rem;color:#7A8CA0;">26AS / Books / Diff</div>
                                        <div style="font-weight:600;color:#1F3864;">
                                            {inr(row['TDS in 26AS'])} &nbsp;/&nbsp; {inr(row['TDS in Books'])} &nbsp;/&nbsp;
                                            <span style="color:{'#C00000' if abs(row['Difference'])>0 else '#166534'};">
                                                {inr(abs(row['Difference']))}
                                            </span>
                                        </div>
                                    </div>
                                </div>
                                {f'<div class="action-box" style="margin-top:10px;">📌 {row["Action"]}</div>' if cat != "Matched" else ""}
                            </div>
                            """, unsafe_allow_html=True)

                # ── Download ──────────────────────────────────────────────────
                st.markdown('<div class="section-title">Download Report</div>', unsafe_allow_html=True)
                excel_bytes = build_excel(df_res, company)
                st.download_button(
                    label="⬇️  Download Full Reconciliation Report (.xlsx)",
                    data=excel_bytes,
                    file_name=f"26AS_Recon_{company.replace(' ','_')}_FY2024-25.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                st.markdown("""
                <div class="disclaimer">
                ⚠️ <b>Disclaimer:</b> This output is indicative only. Matching uses exact TAN+amount, fuzzy vendor name (≥85% similarity), and ±2% amount tolerance.
                Always verify against the actual TRACES 26AS download and audited books before filing returns or responding to notices.
                This tool does not constitute professional tax advice.
                </div>
                """, unsafe_allow_html=True)

            except Exception as e:
                st.error(f"Error processing files: {e}")
                st.info("Please ensure you are uploading the synthetic files generated by this tool, or files in the same format.")

else:
    # Placeholder state
    st.markdown("""
    <div style="text-align:center;padding:60px 20px;color:#7A8CA0;">
        <div style="font-size:3rem;margin-bottom:16px;">⚖️</div>
        <div style="font-family:'DM Serif Display',serif;font-size:1.3rem;color:#1F3864;margin-bottom:8px;">
            Upload your files to begin
        </div>
        <div style="font-size:0.9rem;">
            The reconciliation engine will detect mismatches across 5 categories<br>
            and generate a colour-coded Excel report with action notes.
        </div>
    </div>
    """, unsafe_allow_html=True)

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("""
<div style="text-align:center;font-size:0.78rem;color:#A0AEC0;padding:8px 0;">
    Built by <b>Ankur</b> · Indian Chartered Accountant · AI × Finance · FY 2024-25 Demo
</div>
""", unsafe_allow_html=True)
